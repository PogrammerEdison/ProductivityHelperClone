from datetime import date
import calendar
from openpyxl import load_workbook
import os
from pathlib import Path
import json
import openpyxl
import subprocess
from datetime import datetime
from datetime import timedelta
from PyQt5 import QtGui


class scheduleSelector():
    def __init__(self, jsondict, listWidget, labelList, editingLabel, ssd, table):
        self.excelLocation = os.path.join(Path().absolute(), Path("excelLinearConverter\TemplateExcelforCodeUpdated.xlsx")) # Stores path to excel folder - (In Edison's package)
        self.listWidget = listWidget # Stores list widget - will hold all worksheets
        self.labelList = labelList # Stores all labels - for easy updates
        self.editingLabel = editingLabel # Stores the 'currently selecting for' label
        self.scheduledict = jsondict # Stores the week schedule JSON
        self.ssd = ssd # Stores link to Edisons schedule displayer
        self.currentfocus = calendar.day_name[date.today().weekday()] # Initialised with the current weekday - To be used to store the weekday that the user is currently selecting for 
        self.clickedSheet = self.scheduledict.get(self.currentfocus) # Stores the worksheet that user has last interacted with - intialised with current weekday's assigned sheet
        self.table = table # Store table widget of Edisons displayer

        self.listWidgetUpdater() # Initialises the worksheet list 
        self.labelUpdater() # Initialise labels 
        self.focusUpdater() # Initialise currently selecting for tag 
    
    def refresh(self): # Ensures each time schedule ui entered - ui is set to default 
        self.currentfocus = calendar.day_name[date.today().weekday()]
        self.clickedSheet = self.scheduledict.get(self.currentfocus)
        self.focusUpdater()
        
    def worksheetFinder(self): # Returns list of worksheet names in excel workbook
        wb = load_workbook(filename = self.excelLocation)
        return wb.sheetnames

    def focusUpdater(self): # Changes currently selecting for label to current focus
        self.editingLabel.setText("Currently Selecting For:    " + self.currentfocus)

    def listWidgetUpdater(self): # Updates list widget with all worksheets 
        self.listWidget.clear()
        for item in self.worksheetFinder():
            self.listWidget.addItem(item)

    def listClick(self, item): # Func used when list item is clicked - Refreshes Edison's table 
      worksheet = item.text()
      self.clickedSheet = worksheet
      self.table.setRowCount(0) # Clears table - so new table can be drawn ontop
      self.ssd.displayExcel(self.clickedSheet) # Calls schedule displayer 

    def confirmClick(self): ## Save selection to JSON, Trigger update of worksheet labels.
        self.scheduledict.update({self.currentfocus:self.clickedSheet})
        with open('weekschedule.json', 'w') as fp:
            json.dump(self.scheduledict, fp)
        self.labelUpdater()

        

    
    def excelClick(self): ## Opens excel file to currently displayed worksheet - first needs to be loaded in code and current worksheet set to active - then saved and opened in window
        
        try:
            wb = load_workbook(filename = self.excelLocation)
            wslist = wb.get_sheet_names()
            wb._active_sheet_index = wslist.index(self.clickedSheet)
            wb.save(self.excelLocation)
            subprocess.Popen([self.excelLocation], shell=True)
        except:
            pass

    def weekClick(self, weekday): ## For the Mon - Sun week buttons. # Func run when buttons clicked # Refreshes and displays the currently assigned worksheet to that weekday
        self.currentfocus = weekday
        self.clickedSheet = self.scheduledict.get(weekday)
        self.table.setRowCount(0)
        self.ssd.displayExcel(self.clickedSheet)
        self.focusUpdater()
        
    def labelUpdater(self): ## Used aftear worksheet allocation alterations
        for label in self.labelList:
            label.setText("Selected:    " + self.scheduledict.get(label.property("Weekday")))

        