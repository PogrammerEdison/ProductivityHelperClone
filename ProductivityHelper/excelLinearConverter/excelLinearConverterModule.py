from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def excelToLinear(sheetName):
    workperiods =[]
    wb = load_workbook(filename = 'excelLinearConverter\\TemplateExcelforCodeUpdated.xlsx') # Excel workbook
    sheet=wb[sheetName] # Define worksheet within workbook 
    row_count = sheet.max_row
    column_count = sheet.max_column
    if (column_count % 2) != 0:#there should always be an even number of columns (time, activity)
        column_count = column_count + 1  #if the last column has no activities, it won't count as a column in the excel sheet
    linear = []
    for column in range(1, column_count+1, 2): #for every 2 columns, loop through each row
        for row in range(1, row_count+1): #note that column and row both start with the value 1
            timeCell = sheet.cell(row, column)
            activityCell = sheet.cell(row, column+1) #column will be the time cell and the cell beside it represents the activity
            activityCellvalue = activityType(activityCell.fill.fgColor.rgb)
            linear.append([str(timeCell.value), activityCellvalue])
            if activityCellvalue == "Studying": # cell_range is tuple of tuples (cell, ), (cell, ) - so take index 0
                workperiods.append(str(timeCell.value))
    return linear, workperiods


def activityType(colour):
    if colour == 'FF0070C0': 
        return "Studying"
    else:
        return "Nothing"
